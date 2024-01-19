from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from datetime import datetime
import win32com.client
import os


root = Tk()
root.title("물빛누리 원천징수 GUI")
root.geometry("640x250")
root.resizable(False, False) #최대화가능, 창 크기 값 변경 불가



# 파일지정하기
def file_dest_path():
    file_selected = filedialog.askopenfilename(filetypes=(("xlsx 파일","*.xlsx"),),
                                                  initialdir="C:/") #최초에 설정된 경로)
    if file_selected is None: #사용자가 취소를 누를때
        return
    file_txt_path.delete(0, END)
    file_txt_path.insert(0, file_selected)


# 저장경로 (폴더)
def browse_dest_path():
    folder_selected = filedialog.askdirectory()
    if folder_selected is None: #사용자가 취소를 누를때
        return
    txt_dest_path.delete(0, END)
    txt_dest_path.insert(0, folder_selected)

# 시작버튼 누르면 값 받아오기
def start_tax_data():
    YEAR = str(yearbox.get(1.0,"end-1c"))

    MONTH = int(cmb1_width.get())
    DAY = cmb2_width.get()
    SAVE_RESULT = file_txt_path.get()
    SAVE_FOLDER = txt_dest_path.get()


    wb = openpyxl.load_workbook(SAVE_RESULT, data_only = True)

    ws_personal = wb["인적사항"]
    PERSONAL_DICT = {}

    for row in ws_personal.iter_rows( min_row=2):

        name = row[0].value # 첫 칸 값

        if not name:
            break

        social_id = row[1].value # 주민번호
        address = row[2].value # 주소
        phone = row[3].value # 연락처
        bank = row[4].value # 은행
        banknum = row[5].value # 계좌번호

        #print(name, social_id, address,phone, bank, banknum)

        PERSONAL_DICT[name] = [social_id, address,phone, bank, banknum]


    PAYMENT_DATA = {}
    ws_payment = wb['강사원천징수'] #input

    for row in ws_payment.iter_rows( min_row=2):
        month = row[0].value
        name = row[1].value
        program = row[2].value
        total = row[3].value
        need = row[4].value
        income = row[5].value
        tax1 = row[6].value
        tax2 = row[7].value

        if tax1 is None : 
            continue
        taxtotal = int(tax1) + int(tax2)


        data = [program, total, need, income, tax1, tax2, taxtotal]

        if not PAYMENT_DATA.get(month):
            PAYMENT_DATA[month] = {}

        if not PAYMENT_DATA[month].get(name):
            PAYMENT_DATA[month][name] = []

        PAYMENT_DATA[month][name].append(data)

    # print(PAYMENT_DATA)

    
    ws_reciept = wb['영수증'] # output


    for month, monthly_payment in PAYMENT_DATA.items():
        if month != MONTH:
            continue

        for name, lecturer_payment in monthly_payment.items():

            index = 19

            # lecturer_payment -> [[], [], [], []]
            for data in lecturer_payment:
                print(index)
                print(name)

                ws_reciept["D3"] = YEAR
                ws_reciept["H8"] = name
                ws_reciept["L8"] = PERSONAL_DICT[name][0]
                ws_reciept["H9"] = PERSONAL_DICT[name][1]
                ws_reciept["A34"] = name



                ws_reciept["M31"] = str(YEAR) + "년"
                ws_reciept["N31"] = str(month) + "월"
                ws_reciept["O31"] = str(DAY) + "일"


                # TODO:
                # 이 루프 안에서 계속 다음줄로 넘어가도록 처리해야 함
                # C19 -> C20 이런 식으로...

                ws_reciept["A" + str(index)] = int(YEAR)
                ws_reciept["C" + str(index)] = month
                ws_reciept["D" + str(index)] = DAY
                ws_reciept["E" + str(index)] = int(YEAR)
                ws_reciept["F" + str(index)] = month


                # data = [program, total, need, income, tax1, tax2, taxtotal]


                ws_reciept["G" + str(index)] = data[1]
                ws_reciept["I" + str(index)] = data[2]
                ws_reciept["J" + str(index)] = data[3]

                ws_reciept["K" + str(index)] = "20.00"

                ws_reciept["L" + str(index)] = data[4]
                ws_reciept["M" + str(index)] = data[5]
                ws_reciept["O" + str(index)] = data[6]
        
                
                index = index + 1

            wb.save(SAVE_RESULT)


            # save excel

            output_pdf = os.path.join(SAVE_FOLDER, "{}월{}.pdf".format(month, name))
            # output_pdf = r"C:\Users\semi\Desktop\PythonWorkspace\{}월_{}.pdf".format(month, name)
            

            # create pdf
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False

            wb3 = excel.Workbooks.Open(SAVE_RESULT)

            ws_for_pdf = wb3.Worksheets("영수증")
            ws_for_pdf.Select()

            wb3.ActiveSheet.ExportAsFixedFormat(0, output_pdf)
            

            for cell in ["A19", "A20","A21","A22","A23","A24","A25","A26","A27","A28","A29"]:
                ws_reciept[cell] = ''        
            for cell in ["C19", "C20","C21","C22","C23","C24","C25","C26","C27","C28","C29"]:
                ws_reciept[cell] = ''
            for cell in ["D19", "D20","D21","D22","D23","D24","D25","D26","D27","D28","D29"]:
                ws_reciept[cell] = ''
            for cell in ["E19", "E20","E21","E22","E23","E24","E25","E26","E27","E28","E29"]:
                ws_reciept[cell] = ''
            for cell in ["F19", "F20","F21","F22","F23","F24","F25","F26","F27","F28","F29"]:
                ws_reciept[cell] = ''
            for cell in ["G19", "G20","G21","G22","G23","G24","G25","G26","G27","G28","G29"]:
                ws_reciept[cell] = ''
            for cell in ["I19", "I20","I21","I22","I23","I24","I25","I26","I27","I28","I29"]:
                ws_reciept[cell] = ''
            for cell in ["J19", "J20","J21","J22","J23","J24","J25","J26","J27","J28","J29"]:
                ws_reciept[cell] = ''
            for cell in ["K19", "K20","K21","K22","K23","K24","K25","K26","K27","K28","K29"]:
                ws_reciept[cell] = ''
            for cell in ["L19", "L20","L21","L22","L23","L24","L25","L26","L27","L28","L29"]:
                ws_reciept[cell] = ''
            for cell in ["M19", "M20","M21","M22","M23","M24","M25","M26","M27","M28","M29"]:
                ws_reciept[cell] = ''
            for cell in ["O19", "O20","O21","O22","O23","O24","O25","O26","O27","O28","O29"]:
                ws_reciept[cell] = ''


            wb3.Close(False)
            excel.Quit()
        messagebox.showwarning("완료", "완료")
        excel.Quit()



#시작
def start():
    #날짜입력했는지 확인하기
    if len(cmb1_width.get()) == 0:
        messagebox.showwarning("경고", "월을 선택하세요")
        return
    if len(cmb2_width.get()) == 0:
        messagebox.showwarning("경고", "일을 선택하세요")
        return    

    
    #실적파일을 선택했는지 확인하기
    if len(file_txt_path.get())  == 0:
        messagebox.showwarning("경고", "원청징수 정리 파일을 선택하세요")
        return
    
    #파일저장위치 선택했는지 확인하기
    if len(txt_dest_path.get()) == 0:
        messagebox.showwarning("경고", "저장할 위치를 선택하세요")
        return
    
    start_tax_data()
    
#연도프레임

year_frame = LabelFrame(root, text="몇년")
year_frame.pack()

#연도 텍스트박스
lb1_width = Label(year_frame, text="연도입력")
lb1_width.pack(side="left")


#콤보박스 프레임(파일추가, 선택 삭제)
combobox_frame = LabelFrame(root, text="날짜선택")
combobox_frame.pack()

lb2_width = Label(combobox_frame, text="월선택")
lb2_width.pack(side="left")

#연도 입력 텍스트박스
yearbox = Text(year_frame, width=15, height=1)
yearbox.pack()


#월, 일 선택 콤보박스
cmb1_value = [i for i in range(1,13)]
cmb1_width = ttk.Combobox(combobox_frame, state="readonly", values=cmb1_value, width=10)
cmb1_width.pack(side="left")

cmb2_value = [i for i in range(1,32)]
cmb2_width = ttk.Combobox(combobox_frame, state="readonly", values=cmb2_value, width=10)
cmb2_width.pack(side="left")



# 파일 위치 프레임
file_path_frame = LabelFrame(root, text="원천징수 파일열기")
file_path_frame.pack(fill="both")

file_txt_path = Entry(file_path_frame)
file_txt_path.pack(side="left", fill="x", expand=True)

btn_file_path = Button(file_path_frame, text="파일열기", width=10, command=file_dest_path)
btn_file_path.pack(side="right")


# 저장경로 프레임
path_frame = LabelFrame(root, text="원천징수 영수증 저장위치")
path_frame.pack(fill="both")

txt_dest_path = Entry(path_frame)
txt_dest_path.pack(side="left", fill="x", expand=True)

btn_dest_path = Button(path_frame, text="저장위치", width=10, command=browse_dest_path)
btn_dest_path.pack(side="right")

# 실행 프레임

frame_run = Frame(root)
frame_run.pack(fill="x", padx=5, pady=5)

btn_close = Button(frame_run, padx=5, pady=5, text="닫기", width=12, command=root.quit)
btn_close.pack(side="right", padx=5, pady=5)

btn_start = Button(frame_run, padx=5, pady=5, text="시작", width=12, command=start)
btn_start.pack(side="right", padx=5, pady=5)



root.mainloop()




