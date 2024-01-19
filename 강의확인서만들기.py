import pprint # dictionary 를 보기좋게 출력하려고 쓰는 라이브러리
import openpyxl
from datetime import datetime

raw_data = '교육파일.xlsm'
save_data = '강의확인서.xlsx'

MONTH = 3
EDU = "꾸러기수사대"
# 0. 데이터 통합시트 먼저 읽기
wb = openpyxl.load_workbook(raw_data, read_only=True)
lec_raw_sheet = wb['통합시트']

first_row_number = 2
last_row_number = lec_raw_sheet.max_row


"""
MONTH 데이터만 추출될 거고

{
    'lec_name'{ 
        강사A : [날짜, 시간, 신청기관, 반, 인원],
                [날짜, 시간, 신청기관, 반, 인원]}
        강사b : 
}


"""

PROGRAM_DATA = {}

for row in lec_raw_sheet.iter_rows(min_row=first_row_number, max_row=last_row_number):
    date = row[1].value

    if date != None:
        if date.month != MONTH:
            continue

        lec_name = row[0].value # 강의명
        date = row[1].value.strftime("%Y-%m-%d")  #날짜
        timetable = row[3].value # 시간
        applicant = row[4].value # 신청기관
        division = row[5].value # 반
        amount = row[6].value # 인원
        teacher1 = row[8].value # 강사1
        teacher2 = row[9].value # 강사2

        data = [date, timetable, applicant, division] #월별 데이터 나오는것까지는 성공



        if not lec_name in PROGRAM_DATA:
            PROGRAM_DATA[lec_name] = {}  

        
        for teacher in (teacher1, teacher2):
            if teacher is None:  
                continue
            if teacher is str('-'):
                continue
          
            if not teacher in PROGRAM_DATA[lec_name]:
                PROGRAM_DATA[lec_name][teacher] = []
                
            PROGRAM_DATA[lec_name][teacher].append(data) #data를 붙이면 1개만 나오구 ㅜㅜ


# print("-------------------- PROGRAM_DATA: 출 력 ---------------------------\n")
# pprint.pprint(PROGRAM_DATA, width=1, indent=5)

# #위에서 추출된 PROGRAM_DATA를 가지고 강사별로 강의확인서 인적사항 적기


ws_personal = wb["인적사항"]
PERSONAL_DATA = {}

for row in ws_personal.iter_rows(min_row=2):

    name = row[0].value # 첫 칸 값

    if not name:
        break

    position = row[1].value # 소속직책
    social_id = row[2].value # 생년월일
    address = row[3].value # 주소
    phone = row[4].value # 연락처
    bank = row[5].value # 은행
    banknum = row[6].value # 계좌번호

    # print(name, position, social_id, address,phone, bank, banknum)

    PERSONAL_DATA[name] = [position, social_id, address,phone, bank, banknum]

result = PROGRAM_DATA[EDU]
wb.close()


# EDU에 해당하는 강사의 인적사항만 가져오기

wb2 = openpyxl.load_workbook(save_data, read_only=False)
lec_sheet = wb2['강의확인서']

for teacher_name, lec_detail in result.items():

#인적사항기재하기

    lec_sheet["c5"] = teacher_name #강사명
    lec_sheet["c6"] = PERSONAL_DATA[teacher_name][0] #소속및직위
    lec_sheet["c7"] = PERSONAL_DATA[teacher_name][1]  #생년월일
    lec_sheet["c8"] = PERSONAL_DATA[teacher_name][2]  #주소
    lec_sheet["c9"] = PERSONAL_DATA[teacher_name][3]  #연락처
    lec_sheet["c10"] = PERSONAL_DATA[teacher_name][4]  #은행
    lec_sheet["d10"] = PERSONAL_DATA[teacher_name][5] #계좌번호

    index = 19
    for lec_data in lec_detail:
        lec_sheet["c5"] = teacher_name #강사명

        index = index + 1

    wb2.save("{}월{}.xlsx".format(teacher_name, EDU))
#시트 추가 생성하는걸 자꾸 실패해서 일단 파일을 생성해버렸음
    
#문제점1. 루프가 안돌아감. 덮어쓰기 되어버렸음. -> sheet별로 저장하고싶음
#문제점2. lec_detail을 붙여야 하는데 안됨 ㅠㅠ
    
    





    








            


    

